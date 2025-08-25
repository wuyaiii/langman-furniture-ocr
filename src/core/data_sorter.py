# -*- coding: utf-8 -*-
"""
数据排序处理模块
"""

import re
import os
from pathlib import Path
from ..utils import logger, config_manager

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.error("openpyxl未安装")


class DataSorter:
    """数据排序处理器"""

    def __init__(self, excel_manager):
        self.excel_manager = excel_manager
        self.category_config = config_manager.get_category_config()

        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl是必需的依赖包")

    def sort_excel_data(self):
        """Excel排序处理主函数"""
        try:
            # 读取历史数量数据
            historical_quantities = self.excel_manager.read_historical_quantities()

            # 读取原始数据
            data_groups = self.excel_manager.read_data()
            if not data_groups:
                logger.warning("没有找到可排序的数据")
                return False

            # 解析和分类数据
            parsed_data = self._parse_data_groups(data_groups)
            if not parsed_data:
                logger.warning("没有找到有效的套装数据")
                return False

            # 排序数据
            sorted_data = self._sort_parsed_data(parsed_data)

            # 写入排序结果
            success = self._write_sorted_data(
                sorted_data, historical_quantities)

            if success:
                logger.info("Excel数据排序处理完成")
                return True
            else:
                logger.error("排序数据写入失败")
                return False

        except Exception as e:
            logger.error(f"排序处理失败: {e}")
            return False

    def _parse_data_groups(self, data_groups):
        """解析数据组，提取年份、类别、套装号"""
        parsed_data = []
        seen_combinations = set()  # 用于检测重复的组合

        for group in data_groups:
            title = group['title']
            items = group['items']

            # 解析标题
            parsed_info = self._parse_title(title)
            if not parsed_info:
                continue

            # 创建唯一标识符
            unique_id = f"{parsed_info['year']}_{parsed_info['category']}_{parsed_info['set_number']}_{title}"

            # 检查是否已存在相同的组合
            if unique_id in seen_combinations:
                logger.debug(f"发现重复的组合，跳过: {unique_id}")
                continue

            seen_combinations.add(unique_id)

            parsed_data.append({
                'original_title': title,
                'year': parsed_info['year'],
                'category': parsed_info['category'],
                'set_number': parsed_info['set_number'],
                'items': items,
                'sort_key': (parsed_info['year'], parsed_info['category'], parsed_info['set_number'])
            })

            logger.debug(
                f"解析标题: {title} → 年份:{parsed_info['year']}, 类别:{parsed_info['category']}, 套装号:{parsed_info['set_number']}")

        logger.info(f"解析完成，共 {len(parsed_data)} 个有效套装")
        return parsed_data

    def _parse_title(self, title):
        """解析单个标题，提取年份、类别、套装号"""
        # 提取年份（4位数字）
        year_match = re.search(r'(\d{4})', title)
        if not year_match:
            logger.debug(f"未找到年份: {title}")
            return None

        year = int(year_match.group(1))
        year_end = year_match.end()

        # 提取年份后的部分
        after_year = title[year_end:]

        # 确定类别提取规则
        if '家具套装' in after_year:
            # 提取年份和"家具套装"之间的内容作为类别
            pattern = r'(.+?)家具套装'
            match = re.search(pattern, after_year)
            if match:
                category = match.group(1).strip()
                category = self._clean_category_name(category)
                # 提取套装号
                set_pattern = r'家具套装(\d+)'
                set_match = re.search(set_pattern, after_year)
                set_number = int(set_match.group(1)) if set_match else 1
            else:
                return None

        elif '套装' in after_year:
            # 提取年份和"套装"之间的内容作为类别
            pattern = r'(.+?)套装'
            match = re.search(pattern, after_year)
            if match:
                category = match.group(1).strip()
                category = self._clean_category_name(category)
                # 提取套装号
                set_pattern = r'套装(\d+)'
                set_match = re.search(set_pattern, after_year)
                set_number = int(set_match.group(1)) if set_match else 1
            else:
                return None
        else:
            logger.debug(f"未找到套装关键词: {title}")
            return None

        return {
            'year': year,
            'category': category,
            'set_number': set_number
        }

    def _clean_category_name(self, category):
        """清理类别名称，移除修饰词"""
        if not category:
            return category

        modifiers_to_remove = self.category_config.get(
            'modifiers_to_remove', [])
        cleaned_category = category

        for modifier in modifiers_to_remove:
            cleaned_category = cleaned_category.replace(modifier, '')

        # 去除多余的空格
        cleaned_category = cleaned_category.strip()

        # 如果清理后的类别名为空，返回原始类别名
        if not cleaned_category:
            return category

        # 输出清理日志
        if cleaned_category != category:
            logger.debug(f"类别名清理: '{category}' → '{cleaned_category}'")

        return cleaned_category

    def _sort_parsed_data(self, parsed_data):
        """对解析后的数据进行排序"""
        # 按照年份、类别、套装号排序
        sorted_data = sorted(parsed_data, key=lambda x: x['sort_key'])

        # 按类别分组
        categories = {}
        for item in sorted_data:
            category = item['category']
            if category not in categories:
                categories[category] = []
            categories[category].append(item)

        logger.info(f"找到 {len(categories)} 个类别: {list(categories.keys())}")
        return categories

    def _sort_categories_by_priority(self, categories):
        """按照优先级和内容数量对类别进行排序"""
        priority_order = self.category_config.get('priority_order', [])

        # 创建优先级映射
        priority_map = {category: index for index,
                        category in enumerate(priority_order)}

        # 分离有优先级的类别和无优先级的类别
        prioritized_categories = []
        unprioritized_categories = []

        for category in categories.keys():
            if category in priority_map:
                prioritized_categories.append(category)
            else:
                unprioritized_categories.append(category)

        # 对有优先级的类别按照配置顺序排序
        prioritized_categories.sort(key=lambda x: priority_map[x])

        # 对无优先级的类别按照内容数量（从多到少）排序
        unprioritized_categories.sort(
            key=lambda x: len(categories[x]), reverse=True)

        # 合并排序结果
        sorted_category_names = prioritized_categories + unprioritized_categories

        logger.info(f"类别排序结果:")
        logger.info(
            f"  优先级类别 ({len(prioritized_categories)}个): {prioritized_categories}")
        logger.info(
            f"  其他类别 ({len(unprioritized_categories)}个): {unprioritized_categories}")

        return sorted_category_names

    def _write_sorted_data(self, categories, historical_quantities=None):
        """将排序后的数据写入新工作表"""
        if historical_quantities is None:
            historical_quantities = {}

        try:
            workbook = openpyxl.load_workbook(
                self.excel_manager.excel_file_path)

            # 创建或获取排序结果工作表
            sorted_sheet_name = config_manager.get_env(
                'excel_sorted_sheet_name', '排序结果')

            if sorted_sheet_name in workbook.sheetnames:
                # 如果工作表已存在，删除它
                workbook.remove(workbook[sorted_sheet_name])
                logger.info(f"删除已存在的工作表: {sorted_sheet_name}")

            # 创建新的工作表
            sorted_sheet = workbook.create_sheet(sorted_sheet_name)
            logger.info(f"创建新工作表: {sorted_sheet_name}")

            # 使用自定义排序获取类别顺序
            sorted_category_names = self._sort_categories_by_priority(
                categories)

            # 计算每个类别的列位置
            current_col = 1
            category_positions = {}

            for category in sorted_category_names:
                category_positions[category] = current_col
                item_count = len(categories[category])
                logger.debug(
                    f"类别 '{category}' (共{item_count}项) 分配到列 {get_column_letter(current_col)}-{get_column_letter(current_col+3)}")
                current_col += 5  # 4列数据 + 1列空隙

            # 写入数据
            for category in sorted_category_names:
                items = categories[category]
                start_col = category_positions[category]
                current_row = 1

                logger.debug(f"开始写入类别: {category}")

                # 写入每个套装的数据
                for item in items:
                    # 写入套装标题
                    title_cell = sorted_sheet.cell(
                        row=current_row, column=start_col, value=item['original_title'])
                    title_cell.font = Font(bold=True)
                    title_cell.alignment = Alignment(
                        horizontal='center', vertical='center')

                    # 写入物品名
                    item_row = current_row
                    for item_name in item['items']:
                        item_cell = sorted_sheet.cell(
                            row=item_row, column=start_col + 1, value=item_name)
                        item_cell.font = Font(bold=True)
                        item_cell.alignment = Alignment(
                            horizontal='center', vertical='center')

                        # 检查并写入历史数量数据
                        if item_name in historical_quantities:
                            quantities = historical_quantities[item_name]

                            # 写入第一个数量（C列）
                            if 'quantity1' in quantities:
                                quantity1_cell = sorted_sheet.cell(
                                    row=item_row, column=start_col + 2, value=quantities['quantity1'])
                                quantity1_cell.font = Font(bold=True)
                                quantity1_cell.alignment = Alignment(
                                    horizontal='center', vertical='center')
                                logger.debug(
                                    f"恢复数量数据: {item_name} -> C列: {quantities['quantity1']}")

                            # 写入第二个数量（D列）
                            if 'quantity2' in quantities:
                                quantity2_cell = sorted_sheet.cell(
                                    row=item_row, column=start_col + 3, value=quantities['quantity2'])
                                quantity2_cell.font = Font(bold=True)
                                quantity2_cell.alignment = Alignment(
                                    horizontal='center', vertical='center')
                                logger.debug(
                                    f"恢复数量数据: {item_name} -> D列: {quantities['quantity2']}")

                        item_row += 1

                    # 移动到下一个套装（至少空一行）
                    current_row = max(current_row + 1, item_row + 1)

                logger.debug(f"类别 {category} 写入完成")

            # 应用格式设置
            self._apply_excel_formatting(
                sorted_sheet, categories, category_positions, current_col)

            # 保存文件
            workbook.save(self.excel_manager.excel_file_path)
            workbook.close()

            logger.info(f"排序数据已写入工作表: {sorted_sheet_name}")
            recovered_count = len([k for k in historical_quantities.keys()
                                   if any(item_name == k for items in categories.values()
                                          for item in items for item_name in item['items'])])
            logger.info(f"已恢复 {recovered_count} 个家具的历史数量数据")

            return True

        except Exception as e:
            logger.error(f"写入排序数据失败: {e}")
            return False

    def _apply_excel_formatting(self, sorted_sheet, categories, category_positions, max_col):
        """应用Excel格式设置"""
        try:
            max_row = sorted_sheet.max_row

            # 设置行高为20
            for row in range(1, max_row + 1):
                sorted_sheet.row_dimensions[row].height = 20

            # 设置列宽和格式
            for category in categories.keys():
                start_col = category_positions[category]

                # 第1列：标题列，宽度20.13
                title_col_letter = get_column_letter(start_col)
                sorted_sheet.column_dimensions[title_col_letter].width = 20.13

                # 第2列：物品名列，宽度20.13
                item_col_letter = get_column_letter(start_col + 1)
                sorted_sheet.column_dimensions[item_col_letter].width = 20.13

                # 第3列：剩余列1，宽度5.13
                col3_letter = get_column_letter(start_col + 2)
                sorted_sheet.column_dimensions[col3_letter].width = 5.13

                # 第4列：剩余列2，宽度5.13
                col4_letter = get_column_letter(start_col + 3)
                sorted_sheet.column_dimensions[col4_letter].width = 5.13

                # 第5列：空隙列，宽度1.13，整列标红
                if start_col + 4 <= max_col:
                    gap_col_letter = get_column_letter(start_col + 4)
                    sorted_sheet.column_dimensions[gap_col_letter].width = 1.13

                    # 设置整列背景为红色
                    red_fill = PatternFill(
                        start_color="FF0000", end_color="FF0000", fill_type="solid")

                    for row in range(1, max_row + 1):
                        cell = sorted_sheet.cell(row=row, column=start_col + 4)
                        cell.fill = red_fill

            logger.debug("Excel格式设置完成")

        except Exception as e:
            logger.error(f"应用Excel格式失败: {e}")
