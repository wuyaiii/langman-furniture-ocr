# -*- coding: utf-8 -*-
"""
Excel操作管理模块
"""

import os
import sys
import subprocess
import re
from pathlib import Path
from ..utils import logger, config_manager

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.error("openpyxl未安装")

try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
    logger.warning("xlwings未安装，将仅使用openpyxl")


class ExcelManager:
    """Excel文件管理器"""
    
    def __init__(self):
        self.excel_file_path = None
        self.excel_file_name = None
        self.xlwings_available = XLWINGS_AVAILABLE
        self.openpyxl_available = OPENPYXL_AVAILABLE
        
        if not self.openpyxl_available:
            raise ImportError("openpyxl是必需的依赖包")
        
        self._validate_config()
    
    def _validate_config(self):
        """验证Excel配置"""
        excel_file_path = config_manager.get_env('excel_file_path')
        excel_file_name = config_manager.get_env('excel_file_name')
        
        if excel_file_path:
            self.excel_file_path = excel_file_path
            self.excel_file_name = excel_file_name or os.path.basename(excel_file_path)
            logger.info(f"Excel配置: {self.excel_file_name}")
        else:
            logger.warning("Excel文件路径未配置")
    
    def prepare_excel_file(self):
        """准备Excel文件并用Excel程序打开"""
        if not self.excel_file_path:
            logger.error("Excel文件路径未设置")
            return False
        
        try:
            # 确保目录存在
            file_path = Path(self.excel_file_path)
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 如果文件不存在，创建一个新文件
            if not file_path.exists():
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                
                # 设置工作表名称
                excel_sheet_name = config_manager.get_env('excel_sheet_name')
                if excel_sheet_name:
                    worksheet.title = excel_sheet_name
                
                # 保存文件
                workbook.save(self.excel_file_path)
                workbook.close()
                logger.info(f"创建新的Excel文件: {self.excel_file_path}")
            
            # 用Excel程序打开文件
            self._open_excel_file()
            logger.info(f"Excel文件已准备完成: {self.excel_file_name}")
            return True
            
        except Exception as e:
            logger.error(f"准备Excel文件失败: {e}")
            return False
    
    def _open_excel_file(self):
        """用系统默认程序打开Excel文件"""
        try:
            if sys.platform.startswith('win'):
                os.startfile(self.excel_file_path)
            elif sys.platform.startswith('darwin'):  # macOS
                subprocess.call(['open', self.excel_file_path])
            else:  # Linux
                subprocess.call(['xdg-open', self.excel_file_path])
        except Exception as e:
            logger.error(f"打开Excel文件失败: {e}")
    
    def write_data(self, title, items):
        """写入数据到Excel"""
        if not title and not items:
            logger.warning("没有数据可写入")
            return False
        
        # 确保Excel文件已准备
        if not self.excel_file_path:
            if not self.prepare_excel_file():
                return False
        
        # 优先使用xlwings写入正在运行的Excel
        if self.xlwings_available:
            success = self._write_with_xlwings(title, items)
            if success:
                return True
            logger.warning("xlwings写入失败，尝试使用openpyxl")
        
        # 备用方案：使用openpyxl
        return self._write_with_openpyxl(title, items)
    
    def _write_with_xlwings(self, title, items):
        """使用xlwings写入正在运行的Excel"""
        try:
            # 尝试连接到指定的Excel工作簿
            wb = xw.Book(self.excel_file_name)
            logger.info(f"成功连接到Excel工作簿: {self.excel_file_name}")
            
            # 选择工作表
            excel_sheet_name = config_manager.get_env('excel_sheet_name')
            if excel_sheet_name and excel_sheet_name in [ws.name for ws in wb.sheets]:
                ws = wb.sheets[excel_sheet_name]
            else:
                ws = wb.sheets[0]
            
            # 找到下一个空行
            next_row = self._find_next_empty_row_xlwings(ws)
            current_row = next_row
            
            # 写入标题到A列
            if title:
                ws.range(f'A{current_row}').value = title
                logger.info(f"写入标题到A{current_row}: {title}")
            
            # 写入物品名到B列
            for i, item_name in enumerate(items):
                item_row = current_row + i
                ws.range(f'B{item_row}').value = item_name
                logger.debug(f"写入物品名到B{item_row}: {item_name}")
            
            # 保存文件
            try:
                wb.save()
                logger.info("Excel文件已自动保存")
            except Exception as e:
                logger.warning(f"自动保存失败: {e}")
            
            return True
            
        except Exception as e:
            logger.error(f"xlwings写入失败: {e}")
            return False
    
    def _write_with_openpyxl(self, title, items):
        """使用openpyxl写入Excel文件"""
        try:
            # 检查文件是否被占用
            try:
                with open(self.excel_file_path, 'r+b'):
                    pass
            except PermissionError:
                logger.error(f"Excel文件正在被使用: {self.excel_file_path}")
                return False
            
            # 打开工作簿
            workbook = openpyxl.load_workbook(self.excel_file_path)
            
            # 选择工作表
            excel_sheet_name = config_manager.get_env('excel_sheet_name')
            if excel_sheet_name and excel_sheet_name in workbook.sheetnames:
                worksheet = workbook[excel_sheet_name]
            else:
                worksheet = workbook.active
            
            # 找到下一个空行
            next_row = self._find_next_empty_row_openpyxl(worksheet)
            current_row = next_row
            
            # 写入标题到A列
            if title:
                worksheet.cell(row=current_row, column=1, value=title)
                logger.info(f"写入标题到A{current_row}: {title}")
            
            # 写入物品名到B列
            for i, item_name in enumerate(items):
                item_row = current_row + i
                worksheet.cell(row=item_row, column=2, value=item_name)
                logger.debug(f"写入物品名到B{item_row}: {item_name}")
            
            # 保存并关闭文件
            workbook.save(self.excel_file_path)
            workbook.close()
            logger.info("数据已写入Excel文件")
            return True
            
        except Exception as e:
            logger.error(f"openpyxl写入失败: {e}")
            return False
    
    def _find_next_empty_row_xlwings(self, worksheet):
        """使用xlwings找到下一个空行"""
        try:
            used_range = worksheet.used_range
            if used_range is None:
                return 1
            
            # 检查是否有实际数据
            last_row_with_data = 0
            for row in range(1, used_range.last_cell.row + 1):
                row_has_data = False
                for col in range(1, used_range.last_cell.column + 1):
                    cell_value = worksheet.range(f'{get_column_letter(col)}{row}').value
                    if cell_value is not None and str(cell_value).strip():
                        row_has_data = True
                        break
                
                if row_has_data:
                    last_row_with_data = row
            
            return last_row_with_data + 3 if last_row_with_data > 0 else 1
            
        except Exception as e:
            logger.error(f"查找空行失败: {e}")
            return 1
    
    def _find_next_empty_row_openpyxl(self, worksheet):
        """使用openpyxl找到下一个空行"""
        try:
            max_row = worksheet.max_row if worksheet.max_row > 1 else 1
            max_col = worksheet.max_column if worksheet.max_column > 1 else 1
            
            last_row_with_data = 0
            for row in range(1, max_row + 1):
                row_has_data = False
                for col in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        row_has_data = True
                        break
                
                if row_has_data:
                    last_row_with_data = row
            
            return last_row_with_data + 3 if last_row_with_data > 0 else 1
            
        except Exception as e:
            logger.error(f"查找空行失败: {e}")
            return 1
    
    def read_data(self):
        """读取Excel中的原始数据"""
        if not self.excel_file_path or not os.path.exists(self.excel_file_path):
            logger.error("Excel文件不存在")
            return []
        
        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)
            
            # 选择原始数据工作表
            excel_sheet_name = config_manager.get_env('excel_sheet_name')
            if excel_sheet_name and excel_sheet_name in workbook.sheetnames:
                worksheet = workbook[excel_sheet_name]
            else:
                worksheet = workbook.active
            
            data_groups = []
            current_group = None
            
            # 读取所有数据
            for row in range(1, worksheet.max_row + 1):
                a_value = worksheet.cell(row=row, column=1).value
                b_value = worksheet.cell(row=row, column=2).value
                
                if a_value and str(a_value).strip():
                    # A列有值，这是一个新的标题组
                    if current_group:
                        data_groups.append(current_group)
                    
                    current_group = {
                        'title': str(a_value).strip(),
                        'items': []
                    }
                    
                    # 如果B列也有值，添加到物品列表
                    if b_value and str(b_value).strip():
                        current_group['items'].append(str(b_value).strip())
                
                elif b_value and str(b_value).strip() and current_group:
                    # 只有B列有值，添加到当前组的物品列表
                    current_group['items'].append(str(b_value).strip())
            
            # 添加最后一组
            if current_group:
                data_groups.append(current_group)
            
            workbook.close()
            logger.info(f"读取到 {len(data_groups)} 个数据组")
            return data_groups
            
        except Exception as e:
            logger.error(f"读取Excel数据失败: {e}")
            return []
    
    def read_historical_quantities(self):
        """读取排序结果表中的历史数量数据"""
        if not self.excel_file_path or not os.path.exists(self.excel_file_path):
            logger.error("Excel文件不存在")
            return {}
        
        try:
            workbook = openpyxl.load_workbook(self.excel_file_path)
            sorted_sheet_name = config_manager.get_env('excel_sorted_sheet_name', '排序结果')
            
            # 检查排序结果表是否存在
            if sorted_sheet_name not in workbook.sheetnames:
                logger.info(f"排序结果表 '{sorted_sheet_name}' 不存在，无历史数量数据")
                workbook.close()
                return {}
            
            sorted_sheet = workbook[sorted_sheet_name]
            historical_quantities = {}
            
            # 遍历所有行和列，查找家具名称和对应的数量
            for row in range(1, sorted_sheet.max_row + 1):
                for col in range(1, sorted_sheet.max_column + 1, 5):  # 每5列为一个类别组
                    # B列（col+1）是家具名称，C列（col+2）和D列（col+3）是数量
                    furniture_name_cell = sorted_sheet.cell(row=row, column=col + 1)
                    quantity1_cell = sorted_sheet.cell(row=row, column=col + 2)
                    quantity2_cell = sorted_sheet.cell(row=row, column=col + 3)
                    
                    furniture_name = furniture_name_cell.value
                    quantity1 = quantity1_cell.value
                    quantity2 = quantity2_cell.value
                    
                    # 如果家具名称存在且不为空
                    if furniture_name and str(furniture_name).strip():
                        furniture_key = str(furniture_name).strip()
                        
                        # 保存数量数据（如果存在）
                        quantities = {}
                        if quantity1 is not None and str(quantity1).strip():
                            quantities['quantity1'] = quantity1
                        if quantity2 is not None and str(quantity2).strip():
                            quantities['quantity2'] = quantity2
                        
                        if quantities:  # 只有当存在数量数据时才保存
                            historical_quantities[furniture_key] = quantities
                            logger.debug(f"保存历史数量: {furniture_key} -> {quantities}")
            
            workbook.close()
            logger.info(f"读取到 {len(historical_quantities)} 个家具的历史数量数据")
            return historical_quantities
            
        except Exception as e:
            logger.error(f"读取历史数量数据失败: {e}")
            return {}
